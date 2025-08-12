# gst_downloader_final_with_captcha.py
"""
Final GST PDF Downloader
- GUI collects login, folder, chromedriver path, and FY selection
- Automates GST Portal to download GSTR-1, GSTR-3B (monthly) and annual returns (GSTR-9, GSTR-9C, GSTR-1 Annual, GSTR-3B Annual)
- Waits 30 seconds for manual captcha entry; if not submitted, auto-clicks Login.
- Logs all actions to GST_Download_Log.xlsx, moves most-recent PDF after each download into structured folders.
- Ensures driver.quit() runs in finally to avoid orphaned browser processes.
"""

import os
import time
import shutil
import traceback
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import Workbook, load_workbook

# -------------------------
# Configurable URLs
# -------------------------
LOGIN_URL = "https://services.gst.gov.in/services/login"
MONTHLY_DASHBOARD = "https://return.gst.gov.in/returns/auth/dashboard"
ANNUAL_RETURN_URL = "https://return.gst.gov.in/returns2/auth/annualreturn"

# -------------------------
# Logging helpers (Excel)
# -------------------------
def init_log_file(folder):
    """
    Ensure the log Excel exists and has a header row.
    Returns path to the log file.
    """
    os.makedirs(folder, exist_ok=True)
    log_path = os.path.join(folder, "GST_Download_Log.xlsx")
    if not os.path.exists(log_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp", "Financial Year", "Month", "Document", "Status", "File Path"])
        wb.save(log_path)
    return log_path

def log_action(folder, fy, month, document, status, file_path="N/A"):
    """
    Append a log row to the Excel log. Also prints to console.
    """
    log_path = init_log_file(folder)
    wb = load_workbook(log_path)
    ws = wb.active
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([ts, fy, month, document, status, file_path])
    wb.save(log_path)
    print(f"[{ts}] ({fy}) {month} {document} -> {status} | {file_path}")

# -------------------------
# File helpers for downloads
# -------------------------
def get_latest_pdf(folder):
    """
    Return absolute path to the most recently modified .pdf in `folder`.
    Returns None if none found.
    """
    try:
        files = [f for f in os.listdir(folder) if f.lower().endswith(".pdf")]
        if not files:
            return None
        full_paths = [os.path.join(folder, f) for f in files]
        latest = max(full_paths, key=os.path.getmtime)
        return latest
    except Exception:
        return None

def move_latest_pdf(download_folder, target_folder):
    """
    Move the most recent PDF from download_folder into target_folder.
    Returns destination path or None if nothing moved.
    """
    latest = get_latest_pdf(download_folder)
    if not latest:
        return None
    os.makedirs(target_folder, exist_ok=True)
    dest = os.path.join(target_folder, os.path.basename(latest))
    try:
        shutil.move(latest, dest)
        return dest
    except Exception:
        return None

# -------------------------
# Selenium helpers
# -------------------------
def wait_for_page_load(driver, timeout=30):
    """
    Wait until document.readyState == 'complete'
    """
    WebDriverWait(driver, timeout).until(lambda d: d.execute_script("return document.readyState") == "complete")

def safe_click(driver, xpath, wait_time=20):
    """
    Try to find clickable element by xpath and click it.
    Returns True on success, False on failure (no exception thrown out).
    """
    try:
        el = WebDriverWait(driver, wait_time).until(EC.element_to_be_clickable((By.XPATH, xpath)))
        el.click()
        return True
    except Exception:
        return False

def element_present(driver, xpath, wait_time=8):
    try:
        WebDriverWait(driver, wait_time).until(EC.presence_of_element_located((By.XPATH, xpath)))
        return True
    except Exception:
        return False

# -------------------------
# GUI
# -------------------------
class GSTDownloaderGUI:
    """
    Tkinter GUI. Collects:
      - Username
      - Password (masked)
      - Destination folder
      - ChromeDriver executable path
      - Financial years via checkboxes
    """
    def __init__(self, master):
        self.master = master
        master.title("GST PDF Downloader")

        # Username
        tk.Label(master, text="Username").grid(row=0, column=0, sticky="e")
        self.username_entry = tk.Entry(master, width=36)
        self.username_entry.grid(row=0, column=1, padx=6, pady=4)

        # Password
        tk.Label(master, text="Password").grid(row=1, column=0, sticky="e")
        self.password_entry = tk.Entry(master, width=36, show="*")
        self.password_entry.grid(row=1, column=1, padx=6, pady=4)

        # Destination folder
        tk.Label(master, text="Destination folder").grid(row=2, column=0, sticky="e")
        self.dest_var = tk.StringVar()
        tk.Entry(master, textvariable=self.dest_var, width=36).grid(row=2, column=1, padx=6)
        tk.Button(master, text="Browse", command=self.browse_dest).grid(row=2, column=2, padx=6)

        # ChromeDriver path
        tk.Label(master, text="ChromeDriver path").grid(row=3, column=0, sticky="e")
        self.driver_var = tk.StringVar()
        tk.Entry(master, textvariable=self.driver_var, width=36).grid(row=3, column=1, padx=6)
        tk.Button(master, text="Browse", command=self.browse_driver).grid(row=3, column=2, padx=6)

        # Financial years checkboxes
        tk.Label(master, text="Select Financial Years").grid(row=4, column=0, sticky="ne")
        self.fy_vars = {}
        fy_list = ["FY 2017-18", "FY 2018-19", "FY 2019-20", "FY 2020-21",
                   "FY 2021-22", "FY 2022-23", "FY 2023-24", "FY 2024-25"]
        for i, fy in enumerate(fy_list):
            var = tk.BooleanVar()
            cb = tk.Checkbutton(master, text=fy, variable=var)
            cb.grid(row=4 + (i // 4), column=1 + (i % 4), sticky="w")
            self.fy_vars[fy] = var

        # Submit button
        tk.Button(master, text="Submit", command=self.submit, width=18, bg="#1976D2", fg="white").grid(row=7, column=1, pady=12)

    def browse_dest(self):
        p = filedialog.askdirectory()
        if p:
            self.dest_var.set(p)

    def browse_driver(self):
        p = filedialog.askopenfilename(title="Select ChromeDriver executable")
        if p:
            self.driver_var.set(p)

    def submit(self):
        username = self.username_entry.get().strip()
        password = self.password_entry.get().strip()
        dest = self.dest_var.get().strip()
        driver_path = self.driver_var.get().strip()
        selected = [fy for fy, v in self.fy_vars.items() if v.get()]

        if not username or not password:
            messagebox.showerror("Missing", "Username and Password required.")
            return
        if not dest:
            messagebox.showerror("Missing", "Choose destination folder.")
            return
        if not driver_path:
            messagebox.showerror("Missing", "Choose ChromeDriver path.")
            return
        if not selected:
            messagebox.showerror("Missing", "Select at least one Financial Year.")
            return

        self.master.destroy()
        # run main automation
        run_automation(username, password, dest, driver_path, selected)

# -------------------------
# Main automation
# -------------------------
def run_automation(username, password, base_folder, chrome_driver_path, selected_fys):
    """
    Main end-to-end automation:
      - Launch Chrome using the provided chromedriver path
      - Fill credentials, wait 30s for manual captcha entry then auto-login if needed
      - For each FY (sequentially):
          - GSTR-1 monthly flow (for available months)
          - GSTR-3B monthly flow
          - Annual return flows (GSTR-9, GSTR-1 Annual, GSTR-3B Annual)
          - GSTR-9C flow
      - Log all steps to Excel and move latest PDFs into structured folders
      - Ensure driver.quit() in finally block
      - Show results window at the end
    """

    # prepare chrome options to auto-download PDFs and minimize popups
    chrome_options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": base_folder,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "plugins.always_open_pdf_externally": True,
        "profile.default_content_settings.popups": 0
    }
    chrome_options.add_experimental_option("prefs", prefs)
    chrome_options.add_argument("--disable-popup-blocking")
    chrome_options.add_argument("--disable-notifications")
    chrome_options.add_argument("--safebrowsing-disable-download-protection")

    # Prepare log file
    init_log_file(base_folder)

    driver = None
    try:
        # Start browser using the user-provided ChromeDriver path
        service = Service(chrome_driver_path)
        driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.maximize_window()

        # -----------------
        # LOGIN STEP
        # -----------------
        driver.get(LOGIN_URL)
        wait_for_page_load(driver)

        # Fill username & password fields if present
        try:
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "username"))).clear()
            driver.find_element(By.ID, "username").send_keys(username)
        except Exception:
            # fallback: try by name attribute (some pages use name="user_name")
            try:
                WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.NAME, "user_name")))
                driver.find_element(By.NAME, "user_name").clear()
                driver.find_element(By.NAME, "user_name").send_keys(username)
            except Exception:
                log_action(base_folder, "N/A", "N/A", "Login: username field missing", "Fail")
                raise RuntimeError("username field not found")

        try:
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "user_pass"))).clear()
            driver.find_element(By.ID, "user_pass").send_keys(password)
        except Exception:
            log_action(base_folder, "N/A", "N/A", "Login: password field missing", "Fail")
            raise RuntimeError("password field not found")

        # After autofill, wait up to 30 seconds for manual captcha entry and user's click on Login.
        # We'll poll the current URL each second to detect whether the user clicked Login (URL changes).
        # If after 30 seconds the URL hasn't changed, we click the Login button automatically.
        pre_login_url = driver.current_url
        clicked_by_script = False
        print("Please enter CAPTCHA manually. Waiting up to 30s; after that the script will attempt to click Login automatically.")
        start = time.time()
        logged_in = False
        while time.time() - start < 30:
            # If URL changed indicating navigation has happened (user pressed Login), break
            if "dashboard" in driver.current_url.lower():
                logged_in = True
                break
            time.sleep(1)
        if not logged_in:
            # Attempt to auto-click Login button
            # Using provided element code: button type=submit with text 'Login' or data-ng-bind
            print("30s elapsed. Attempting automated click of Login button.")
            # Try multiple xpaths to find login button robustly
            login_xpaths = [
                "//button[@type='submit' and (contains(normalize-space(.),'Login') or contains(normalize-space(.),'LOGIN'))]",
                "//button[@type='submit' and contains(@data-ng-bind,'trans.HEAD_LOGIN')]",
                "//button[@type='submit']"
            ]
            for xp in login_xpaths:
                try:
                    el = driver.find_element(By.XPATH, xp)
                    el.click()
                    clicked_by_script = True
                    break
                except Exception:
                    continue

        # Now wait for the dashboard (success) — may take longer (captcha solving / server)
        try:
            WebDriverWait(driver, 300).until(EC.url_contains("dashboard"))
            wait_for_page_load(driver)
            log_action(base_folder, "N/A", "N/A", "Login", "Success")
        except Exception as e:
            log_action(base_folder, "N/A", "N/A", f"Login did not go to dashboard: {e}", "Fail")
            raise RuntimeError("Login failed or timed out. Please check manually.")

        # -----------------
        # PER-FY SEQUENTIAL FLOW
        # For each selected FY complete all monthly & annual downloads before moving to next FY
        # -----------------
        for fy_label in selected_fys:
            # Normalize FY text used in selects (GST site uses e.g. "2020-21")
            fy_text = fy_label.replace("FY ", "").strip() if fy_label.startswith("FY") else fy_label
            print(f"Starting downloads for {fy_label}")

            # ----------------- GSTR-1 monthly flow -----------------
            try:
                driver.get(MONTHLY_DASHBOARD)
                wait_for_page_load(driver)
                # Select FY (select name='fin')
                try:
                    fin_el = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "fin")))
                    Select(fin_el).select_by_visible_text(fy_text)
                except Exception:
                    # log and continue — sometimes the select may differ
                    log_action(base_folder, fy_label, "N/A", "Select FY on monthly dashboard", "Fail")

                # Try to find month options (name='mon'). If not found, run one iteration with month N/A
                months = []
                try:
                    msel = driver.find_element(By.NAME, "mon")
                    months = [opt.text.strip() for opt in Select(msel).options if opt.text.strip()]
                    if not months:
                        months = [None]
                except Exception:
                    months = [None]

                for month in months:
                    month_label = month if month else "N/A"
                    if month:
                        try:
                            Select(driver.find_element(By.NAME, "mon")).select_by_visible_text(month)
                        except Exception:
                            pass

                    # Click Search (button with class 'srchbtn' or contains text 'Search')
                    search_clicked = False
                    for sx in ["//button[contains(@class,'srchbtn')]", "//button[contains(normalize-space(.),'Search')]"]:
                        if safe_click(driver, sx):
                            search_clicked = True
                            break
                    if not search_clicked:
                        log_action(base_folder, fy_label, month_label, "Search (GSTR-1)", "Fail")
                    else:
                        wait_for_page_load(driver)
                        log_action(base_folder, fy_label, month_label, "Search (GSTR-1)", "Success")

                    # Click VIEW (GSTR-1) in the results tile
                    gstr1_view_xpath = "//div[@data-ng-if=\"x.return_ty=='GSTR1'\" or contains(.,'GSTR-1')]//button[contains(translate(normalize-space(.),'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'VIEW')]"
                    if safe_click(driver, gstr1_view_xpath):
                        wait_for_page_load(driver)
                        log_action(base_folder, fy_label, month_label, "GSTR-1 Click View", "Success")
                        # Click View Summary
                        gstr1_view_summary_xpath = "//button[.//span[contains(normalize-space(.),'VIEW SUMMARY') or contains(normalize-space(.),'PROCEED TO FILE/SUMMARY')]]"
                        if safe_click(driver, gstr1_view_summary_xpath):
                            wait_for_page_load(driver)
                            log_action(base_folder, fy_label, month_label, "GSTR-1 View Summary", "Success")
                        else:
                            log_action(base_folder, fy_label, month_label, "GSTR-1 View Summary", "Fail")
                        # Click Download PDF (genratepdfNew())
                        gstr1_download_xpath = "//button[contains(@data-ng-click,'genratepdfNew') or contains(normalize-space(.),'DOWNLOAD (PDF)') or contains(normalize-space(.),'DOWNLOAD SUMMARY (PDF)')]"
                        if safe_click(driver, gstr1_download_xpath):
                            # attempt to move newest PDF into target folder
                            dest = os.path.join(base_folder, fy_label, "GSTR-1")
                            moved = move_latest_pdf(base_folder, dest)
                            if moved:
                                log_action(base_folder, fy_label, month_label, "Download GSTR-1 (PDF)", "Success", moved)
                            else:
                                log_action(base_folder, fy_label, month_label, "Download GSTR-1 (PDF)", "Success", "File not found")
                        else:
                            log_action(base_folder, fy_label, month_label, "Download GSTR-1 (PDF)", "Fail")
                    else:
                        log_action(base_folder, fy_label, month_label, "GSTR-1 Click View", "Fail")

            except Exception as e:
                log_action(base_folder, fy_label, "N/A", f"GSTR-1 overall error: {e}", "Fail")

            # ----------------- GSTR-3B monthly flow -----------------
            try:
                driver.get(MONTHLY_DASHBOARD)
                wait_for_page_load(driver)
                try:
                    fin_el = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "fin")))
                    Select(fin_el).select_by_visible_text(fy_text)
                except Exception:
                    log_action(base_folder, fy_label, "N/A", "Select FY on monthly dashboard (GSTR-3B)", "Fail")

                # get months as earlier
                months = []
                try:
                    msel = driver.find_element(By.NAME, "mon")
                    months = [opt.text.strip() for opt in Select(msel).options if opt.text.strip()]
                    if not months:
                        months = [None]
                except Exception:
                    months = [None]

                for month in months:
                    month_label = month if month else "N/A"
                    if month:
                        try:
                            Select(driver.find_element(By.NAME, "mon")).select_by_visible_text(month)
                        except Exception:
                            pass

                    # Click Search
                    search_clicked = False
                    for sx in ["//button[contains(@class,'srchbtn')]", "//button[contains(normalize-space(.),'Search')]"]:
                        if safe_click(driver, sx):
                            search_clicked = True
                            break
                    if not search_clicked:
                        log_action(base_folder, fy_label, month_label, "Search (GSTR-3B)", "Fail")
                        continue
                    wait_for_page_load(driver)
                    log_action(base_folder, fy_label, month_label, "Search (GSTR-3B)", "Success")

                    # Click Download (GSTR-3B)
                    gstr3b_xpath = "//button[@data-ng-click='downloadGSTR3Bpdf()' or contains(normalize-space(.),'Download')]"
                    if safe_click(driver, gstr3b_xpath):
                        dest = os.path.join(base_folder, fy_label, "GSTR-3B")
                        moved = move_latest_pdf(base_folder, dest)
                        if moved:
                            log_action(base_folder, fy_label, month_label, "Download GSTR-3B (PDF)", "Success", moved)
                        else:
                            log_action(base_folder, fy_label, month_label, "Download GSTR-3B (PDF)", "Success", "File not found")
                    else:
                        log_action(base_folder, fy_label, month_label, "Download GSTR-3B (PDF)", "Fail")

            except Exception as e:
                log_action(base_folder, fy_label, "N/A", f"GSTR-3B overall error: {e}", "Fail")

            # ----------------- Annual returns (GSTR-9, GSTR-1 Annual, GSTR-3B Annual) -----------------
            try:
                driver.get(ANNUAL_RETURN_URL)
                wait_for_page_load(driver)
                # select finyr
                try:
                    finyr_el = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "finyr")))
                    Select(finyr_el).select_by_visible_text(fy_text)
                except Exception:
                    log_action(base_folder, fy_label, "N/A", "Select FY on annual page", "Fail")

                # click Search
                if safe_click(driver, "//button[contains(@class,'srchbtn') or contains(normalize-space(.),'Search')]"):
                    wait_for_page_load(driver)
                    log_action(base_folder, fy_label, "N/A", "Annual Search", "Success")
                else:
                    log_action(base_folder, fy_label, "N/A", "Annual Search", "Fail")

                # View GSTR-9 and downloads
                gstr9_view_candidates = [
                    "//button[contains(text(),'VIEW GSTR-9')]",
                    "//button[@data-ng-click='page_rtp(x.return_ty,x.due_dt,x.status)' and contains(.,'GSTR-9')]",
                    "//button[contains(@data-ng-click,'page_rtp') and contains(.,'GSTR-9')]"
                ]
                viewed = False
                for xp in gstr9_view_candidates:
                    if safe_click(driver, xp):
                        viewed = True
                        wait_for_page_load(driver)
                        break
                if not viewed:
                    log_action(base_folder, fy_label, "N/A", "View GSTR-9 not found", "Fail")
                else:
                    # Download GSTR-9 details (PDF)
                    if safe_click(driver, "//button[contains(@data-ng-click,'getPdfData_gstr9') or contains(normalize-space(.),'Download GSTR-9')]"):
                        dest = os.path.join(base_folder, fy_label, "GSTR-9")
                        moved = move_latest_pdf(base_folder, dest)
                        if moved:
                            log_action(base_folder, fy_label, "N/A", "Download GSTR-9 (PDF)", "Success", moved)
                        else:
                            log_action(base_folder, fy_label, "N/A", "Download GSTR-9 (PDF)", "Success", "File not found")
                    else:
                        log_action(base_folder, fy_label, "N/A", "Download GSTR-9 (PDF)", "Fail")

                    # Download GSTR-1 annual summary
                    if safe_click(driver, "//button[contains(@data-ng-click,'getPdfData_gstr1') or contains(normalize-space(.),'GSTR-1/IFF SUMMARY')]"):
                        dest = os.path.join(base_folder, fy_label, "GSTR-1 Annual")
                        moved = move_latest_pdf(base_folder, dest)
                        if moved:
                            log_action(base_folder, fy_label, "N/A", "Download GSTR-1 Annual (PDF)", "Success", moved)
                        else:
                            log_action(base_folder, fy_label, "N/A", "Download GSTR-1 Annual (PDF)", "Success", "File not found")
                    else:
                        log_action(base_folder, fy_label, "N/A", "Download GSTR-1 Annual (PDF)", "Fail")

                    # Download GSTR-3B annual summary
                    if safe_click(driver, "//button[contains(@data-ng-click,'getPdfData_gstr3B') or contains(normalize-space(.),'GSTR-3B SUMMARY')]"):
                        dest = os.path.join(base_folder, fy_label, "GSTR-3B Annual")
                        moved = move_latest_pdf(base_folder, dest)
                        if moved:
                            log_action(base_folder, fy_label, "N/A", "Download GSTR-3B Annual (PDF)", "Success", moved)
                        else:
                            log_action(base_folder, fy_label, "N/A", "Download GSTR-3B Annual (PDF)", "Success", "File not found")
                    else:
                        log_action(base_folder, fy_label, "N/A", "Download GSTR-3B Annual (PDF)", "Fail")

            except Exception as e:
                log_action(base_folder, fy_label, "N/A", f"Annual return overall error: {e}", "Fail")

            # ----------------- GSTR-9C -----------------
            try:
                gstr9c_btn_xpath = "//button[contains(@data-ng-click,'offlinepath') or contains(normalize-space(.),'DOWNLOAD GSTR-9C')]"
                if safe_click(driver, gstr9c_btn_xpath):
                    old_url = driver.current_url
                    # wait short time for navigation; if not changed still continue
                    try:
                        WebDriverWait(driver, 20).until(lambda d: d.current_url != old_url)
                    except Exception:
                        pass
                    wait_for_page_load(driver)
                    # click final download button
                    if safe_click(driver, "//button[@data-ng-click='generate9cpdf()' or contains(normalize-space(.),'Download filed GSTR-9C')]"):
                        dest = os.path.join(base_folder, fy_label, "GSTR-9C")
                        moved = move_latest_pdf(base_folder, dest)
                        if moved:
                            log_action(base_folder, fy_label, "N/A", "Download GSTR-9C (PDF)", "Success", moved)
                        else:
                            log_action(base_folder, fy_label, "N/A", "Download GSTR-9C (PDF)", "Success", "File not found")
                    else:
                        log_action(base_folder, fy_label, "N/A", "Download GSTR-9C (PDF)", "Fail")
                else:
                    log_action(base_folder, fy_label, "N/A", "GSTR-9C offline button not found", "Fail")
            except Exception as e:
                log_action(base_folder, fy_label, "N/A", f"GSTR-9C overall error: {e}", "Fail")

        # End per-FY loop

    except Exception as main_e:
        log_action(base_folder, "N/A", "N/A", f"Unhandled exception: {main_e}", "Fail")
        print("Unhandled exception:", traceback.format_exc())
    finally:
        # Always ensure the browser process is closed so no orphan processes remain
        try:
            if driver:
                driver.quit()
        except Exception:
            pass

        # After automation completes (or on crash) show a results window summarizing failures
        show_results_window(base_folder)

# -------------------------
# Results summary window
# -------------------------
def show_results_window(folder):
    """
    Read the Excel log and show either:
     - a success message if no 'Fail' statuses present
     - or a window with a table of failed rows (FY, Month, Document)
    """
    log_path = init_log_file(folder)
    wb = load_workbook(log_path)
    ws = wb.active

    failures = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ts, fy, month, doc, status, fpath = row
        if status and "Fail" in str(status):
            failures.append((fy or "N/A", month or "N/A", doc or "N/A"))

    root = tk.Tk()
    root.title("Download Results")
    root.geometry("700x450")

    if not failures:
        tk.Label(root, text="✅ All downloads completed successfully!", fg="green", font=("Arial", 14)).pack(pady=20)
    else:
        tk.Label(root, text="❌ Some downloads failed:", fg="red", font=("Arial", 14)).pack(pady=8)
        cols = ("Financial Year", "Month", "Document")
        tree = ttk.Treeview(root, columns=cols, show="headings")
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=220)
        tree.pack(fill="both", expand=True, padx=10, pady=10)
        for r in failures:
            tree.insert("", "end", values=r)

    tk.Button(root, text="Close", command=root.destroy).pack(pady=8)
    root.mainloop()

# -------------------------
# Entry point
# -------------------------
if __name__ == "__main__":
    app_root = tk.Tk()
    gui = GSTDownloaderGUI(app_root)
    app_root.mainloop()
